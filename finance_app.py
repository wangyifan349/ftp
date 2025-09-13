# finance_app.py
"""
财务管理程序 

这是一个用于本地或受信环境的单文件财务管理应用（Python + Flask），功能包括用户注册/登录、记录与管理个人交易、每月统计、图表展示与 Excel 导出。后端使用 sqlite3 保存数据，前端使用 Bootstrap 5 提供界面，Chart.js 绘制每月收入与支出柱状图，openpyxl 用于生成可下载的 Excel 文件。密码使用 bcrypt 哈希存储。
主要特性
- 单文件部署：整个应用包含在一个文件 finance_app.py 中，便于本地运行与快速演示。
- 用户认证：支持注册与登录（用户名唯一、密码以 bcrypt 哈希保存）。会话在前端使用 localStorage 简单保存当前用户信息（适合本地/受信环境，不适合生产级会话管理）。
- 多用户隔离：每条交易记录关联到创建者的 user_id，API 在读取/修改/删除交易时通过 user_id 验证权限，确保每个人管理自己的财务数据。
- 交易管理：支持添加、编辑、删除交易。交易字段包含描述、金额、类型（income 或 expense）、分类与日期（ISO 格式）。
- 过滤与查询：可以按用户、分类、类型与日期范围拉取交易数据；按月过滤支持通过 start_date/end_date 查询。
- 每月汇总：后端提供按月聚合的收入与支出汇总（返回 income、expense、balance），用于表格与图表展示。
- 图表展示：前端使用 Chart.js 绘制柱状图，分别显示收入和支出两组数据，支持鼠标提示与清晰对比。
- Excel 导出：提供 /api/export 和 /api/export_excel 两个路由，使用 openpyxl 生成包含 “Transactions” 与 “Monthly Summary” 两个工作表的 .xlsx 文件并触发下载。
- 界面主题：页面采用淡绿色与淡淡金色的卡片主题，整体背景为淡红色，按钮与文本色彩配合以提高可读性（仅视觉样式修改，不影响功能）。
实现与技术细节
- 后端框架：Flask（单进程示例，app.run(debug=False)）。
- 数据库：sqlite3（文件 finance.db）。首次运行时自动创建 users 与 transactions 表，并插入示例用户 demo/demo（密码以 bcrypt 哈希保存）。
- 密码安全：使用 bcrypt.gensalt() + bcrypt.hashpw() 存储密码为二进制哈希；登录时使用 bcrypt.checkpw() 验证。
- SQL 使用：直接以原生 SQL 语句操作 sqlite，未使用 ORM。数据库连接在 Flask 应用上下文中通过 g 缓存，并在上下文结束时关闭。
- API 路由（主要）：
  - POST /api/register — 注册新用户，返回 user_id 与 username。
  - POST /api/login — 登录验证，返回 user_id 与 username。
  - POST /api/add_transaction — 添加交易，需要 JSON 中包含 user_id。
  - POST /api/edit_transaction — 编辑交易，需提供 user_id 与交易 id，且交易必须属于该用户。
  - POST /api/delete_transaction — 删除交易，需提供 user_id 与交易 id，且交易必须属于该用户。
  - GET /api/get_transactions — 获取交易列表，支持 user_id、start_date、end_date、category、type 过滤。
  - GET /api/monthly_summary — 返回按月的收入/支出/结余聚合，用于图表与表格。
  - GET /api/export 和 /api/export_excel — 导出当前用户所有交易与月度汇总为 Excel 文件。
- 前端实现：
  - 单页面 HTML 返回（由 Flask 的 index 路由直接以三联引号字符串返回完整页面）。
  - 使用 localStorage 存储当前登录用户（键名 currentUser），所有对 API 的调用通过在请求体或查询参数中带上 user_id 实现“会话”识别。
  - 交易表、月度表格与图表都在前端渲染；图表配置已确保同时显示收入与支出两条 dataset（绿色表示收入、红色表示支出）。
  - 简易快速添加与高级添加两种方式：快速在顶部填写描述/金额/分类并一键新增收入或支出；高级添加使用 prompt 支持自定义日期。
已知限制与安全提示（重要）
- 仅适合本地或受信网络环境做演示或个人使用。当前实现并未使用服务器端会话管理、CSRF 保护、表单/输入的严格服务端验证或 HTTPS 强制。
- 前端基于 localStorage 存储 user_id，易被客户端篡改；仅用作方便演示，不能作为真实安全认证会话。
- raw SQL 语句直接拼接参数均通过参数化查询传递，但仍建议在生产中使用更严格输入校验与 ORM 以降低开发错误风险。
- 若要在公网部署：务必启用 HTTPS、引入安全会话（服务器端签名 cookie）、CSRF 保护、加强密码策略与速率限制、对导出与文件处理添加权限与审计等。

如何运行（简要）
1. 安装依赖：pip install flask bcrypt openpyxl
2. 将完整文件保存为 finance_app.py，运行：python finance_app.py
3. 在浏览器访问 http://127.0.0.1:5000 ，使用示例账户 demo/demo 或自行注册新用户。
4. 所有数据保存在本地文件 finance.db 中。

总结
这是一个面向本地与受信场景的轻量级单文件财务管理应用，覆盖用户认证、交易管理、月度汇总、图表展示与 Excel 导出。视觉主题已根据要求调整为淡绿色/淡淡金色卡片和淡红色背景，Chart.js 图表明确区分收入与支出，且每个用户的数据被隔离，适合个人或小范围演示与本地使用。
"""
from flask import Flask, g, request, jsonify, send_file
import sqlite3, os
from datetime import datetime
from io import BytesIO
import openpyxl
import bcrypt
DB_PATH = 'finance.db'
app = Flask(__name__)
# ---------- Database helpers (sqlite3 raw SQL) ----------
def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        need_init = not os.path.exists(DB_PATH)
        db = g._database = sqlite3.connect(DB_PATH, check_same_thread=False)
        db.row_factory = sqlite3.Row
        if need_init:
            init_db(db)
    return db
def init_db(db):
    cur = db.cursor()
    # users table, password stored as bcrypt hash
    cur.execute('''
    CREATE TABLE users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash BLOB NOT NULL
    )
    ''')
    # transactions table
    cur.execute('''
    CREATE TABLE transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        description TEXT NOT NULL,
        amount REAL NOT NULL,
        type TEXT NOT NULL, -- income or expense
        category TEXT,
        date TEXT NOT NULL,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    )
    ''')
    # insert example user demo/demo (hashed)
    pwd = b"demo"
    salt = bcrypt.gensalt()
    ph = bcrypt.hashpw(pwd, salt)
    cur.execute("INSERT INTO users (username, password_hash) VALUES (?, ?)", ("demo", ph))
    db.commit()
@app.teardown_appcontext
def close_db(exc):
    db = getattr(g, '_database', None)
    if db:
        db.close()
# ---------- User helpers ----------
def get_user_by_username(username):
    cur = get_db().execute("SELECT id, username, password_hash FROM users WHERE username = ?", (username,))
    return cur.fetchone()
def get_user_by_id(uid):
    cur = get_db().execute("SELECT id, username FROM users WHERE id = ?", (uid,))
    return cur.fetchone()
# ---------- API: register / login (bcrypt) ----------
@app.route('/api/register', methods=['POST'])
def api_register():
    data = request.json or {}
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    if not username or not password:
        return jsonify({"error":"username and password required"}), 400
    if get_user_by_username(username):
        return jsonify({"error":"username exists"}), 400
    pw_bytes = password.encode('utf-8')
    salt = bcrypt.gensalt()
    pw_hash = bcrypt.hashpw(pw_bytes, salt)
    cur = get_db().cursor()
    cur.execute("INSERT INTO users (username, password_hash) VALUES (?, ?)", (username, pw_hash))
    get_db().commit()
    return jsonify({"message":"registered","user_id": cur.lastrowid, "username": username}), 201
@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.json or {}
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    user = get_user_by_username(username)
    if not user:
        return jsonify({"error":"invalid username or password"}), 401
    pw_bytes = password.encode('utf-8')
    stored = user['password_hash']
    try:
        ok = bcrypt.checkpw(pw_bytes, stored)
    except Exception:
        ok = False
    if not ok:
        return jsonify({"error":"invalid username or password"}), 401
    return jsonify({"message":"login successful","user_id": user['id'], "username": user['username']}), 200
# ---------- Helper: require user by id in request body ----------
def require_user_from(data):
    uid = data.get('user_id')
    if not uid:
        return None, (jsonify({"error":"missing user_id"}), 401)
    user = get_user_by_id(uid)
    if not user:
        return None, (jsonify({"error":"invalid user_id"}), 401)
    return user, None
# ---------- Transactions API (raw SQL) ----------
@app.route('/api/add_transaction', methods=['POST'])
def api_add_transaction():
    data = request.json or {}
    user, err = require_user_from(data)
    if err: return err
    description = (data.get('description') or '').strip()
    try:
        amount = float(data.get('amount'))
    except:
        return jsonify({"error":"invalid amount"}), 400
    ttype = data.get('type')
    if ttype not in ('income','expense'):
        return jsonify({"error":"type must be 'income' or 'expense'"}), 400
    category = (data.get('category') or '').strip() or None
    date_str = data.get('date') or datetime.utcnow().isoformat()
    cur = get_db().cursor()
    cur.execute("""INSERT INTO transactions (user_id, description, amount, type, category, date)
                   VALUES (?, ?, ?, ?, ?, ?)""", (user['id'], description, amount, ttype, category, date_str))
    get_db().commit()
    return jsonify({"message":"added","transaction_id": cur.lastrowid}), 201
@app.route('/api/edit_transaction', methods=['POST'])
def api_edit_transaction():
    data = request.json or {}
    user, err = require_user_from(data)
    if err: return err
    tx_id = data.get('id')
    if not tx_id:
        return jsonify({"error":"missing transaction id"}), 400
    cur = get_db().execute("SELECT * FROM transactions WHERE id=? AND user_id=?", (tx_id, user['id']))
    tx = cur.fetchone()
    if not tx:
        return jsonify({"error":"transaction not found or no permission"}), 404
    description = data.get('description', tx['description'])
    try:
        amount = float(data.get('amount', tx['amount']))
    except:
        return jsonify({"error":"invalid amount"}), 400
    ttype = data.get('type', tx['type'])
    if ttype not in ('income','expense'):
        return jsonify({"error":"type must be 'income' or 'expense'"}), 400
    category = data.get('category', tx['category'])
    date_str = data.get('date', tx['date'])
    get_db().execute("""UPDATE transactions SET description=?, amount=?, type=?, category=?, date=? WHERE id=? AND user_id=?""",
                     (description, amount, ttype, category, date_str, tx_id, user['id']))
    get_db().commit()
    return jsonify({"message":"updated"}), 200
@app.route('/api/delete_transaction', methods=['POST'])
def api_delete_transaction():
    data = request.json or {}
    user, err = require_user_from(data)
    if err: return err
    tx_id = data.get('id')
    if not tx_id:
        return jsonify({"error":"missing transaction id"}), 400
    cur = get_db().cursor()
    cur.execute("DELETE FROM transactions WHERE id=? AND user_id=?", (tx_id, user['id']))
    get_db().commit()
    if cur.rowcount == 0:
        return jsonify({"error":"transaction not found or no permission"}), 404
    return jsonify({"message":"deleted"}), 200
# ---------- Get transactions (with filters) ----------
@app.route('/api/get_transactions', methods=['GET'])
def api_get_transactions():
    user_id = request.args.get('user_id', type=int)
    if not user_id:
        return jsonify({"error":"please provide user_id"}), 400
    if not get_user_by_id(user_id):
        return jsonify({"error":"invalid user_id"}), 401
    q = "SELECT * FROM transactions WHERE user_id = ?"
    params = [user_id]
    start = request.args.get('start_date')
    end = request.args.get('end_date')
    if start:
        q += " AND date >= ?"; params.append(start)
    if end:
        q += " AND date <= ?"; params.append(end)
    category = request.args.get('category')
    if category:
        q += " AND category = ?"; params.append(category)
    ttype = request.args.get('type')
    if ttype in ('income','expense'):
        q += " AND type = ?"; params.append(ttype)
    q += " ORDER BY date DESC"
    cur = get_db().execute(q, params)
    rows = cur.fetchall()
    txs = []
    for r in rows:
        txs.append({"id": r['id'], "description": r['description'], "amount": r['amount'],
                    "type": r['type'], "category": r['category'], "date": r['date']})
    return jsonify(txs), 200
# ---------- Monthly summary (for charts) ----------
@app.route('/api/monthly_summary', methods=['GET'])
def api_monthly_summary():
    user_id = request.args.get('user_id', type=int)
    if not user_id:
        return jsonify({"error":"please provide user_id"}), 400
    if not get_user_by_id(user_id):
        return jsonify({"error":"invalid user_id"}), 401
    cur = get_db().execute("""
      SELECT substr(date,1,7) as month,
             SUM(CASE WHEN type='income' THEN amount ELSE 0 END) as income,
             SUM(CASE WHEN type='expense' THEN amount ELSE 0 END) as expense
      FROM transactions
      WHERE user_id=?
      GROUP BY month
      ORDER BY month ASC
    """, (user_id,))
    rows = cur.fetchall()
    res = []
    for r in rows:
        income = r['income'] or 0
        expense = r['expense'] or 0
        res.append({"month": r['month'], "income": income, "expense": expense, "balance": income - expense})
    return jsonify(res), 200
# ---------- Export Excel (existing route, kept) ----------
@app.route('/api/export_excel', methods=['GET'])
def api_export_excel():
    user_id = request.args.get('user_id', type=int)
    if not user_id:
        return jsonify({"error":"please provide user_id"}), 400
    user = get_user_by_id(user_id)
    if not user:
        return jsonify({"error":"invalid user_id"}), 401
    cur = get_db().execute("SELECT * FROM transactions WHERE user_id=? ORDER BY date DESC", (user_id,))
    rows = cur.fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(["ID","date","description","type","category","amount"])
    for r in rows:
        ws.append([r['id'], r['date'], r['description'], r['type'], r['category'] or '', r['amount']])
    # monthly summary sheet
    cur2 = get_db().execute("""
      SELECT substr(date,1,7) as month,
             SUM(CASE WHEN type='income' THEN amount ELSE 0 END) as income,
             SUM(CASE WHEN type='expense' THEN amount ELSE 0 END) as expense
      FROM transactions WHERE user_id=? GROUP BY month ORDER BY month ASC
    """, (user_id,))
    ws2 = wb.create_sheet("Monthly Summary")
    ws2.append(["month","income","expense","balance"])
    for r in cur2.fetchall():
        income = r['income'] or 0
        expense = r['expense'] or 0
        ws2.append([r['month'], income, expense, income - expense])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"finance_{user['username']}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
    return send_file(bio, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ---------- New Export route (English named) ----------
@app.route('/api/export', methods=['GET'])
def api_export():
    # same implementation as api_export_excel, named 'export'
    user_id = request.args.get('user_id', type=int)
    if not user_id:
        return jsonify({"error":"please provide user_id"}), 400
    user = get_user_by_id(user_id)
    if not user:
        return jsonify({"error":"invalid user_id"}), 401
    cur = get_db().execute("SELECT * FROM transactions WHERE user_id=? ORDER BY date DESC", (user_id,))
    rows = cur.fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(["ID","date","description","type","category","amount"])
    for r in rows:
        ws.append([r['id'], r['date'], r['description'], r['type'], r['category'] or '', r['amount']])
    cur2 = get_db().execute("""
      SELECT substr(date,1,7) as month,
             SUM(CASE WHEN type='income' THEN amount ELSE 0 END) as income,
             SUM(CASE WHEN type='expense' THEN amount ELSE 0 END) as expense
      FROM transactions WHERE user_id=? GROUP BY month ORDER BY month ASC
    """, (user_id,))
    ws2 = wb.create_sheet("Monthly_Summary")
    ws2.append(["month","income","expense","balance"])
    for r in cur2.fetchall():
        income = r['income'] or 0
        expense = r['expense'] or 0
        ws2.append([r['month'], income, expense, income - expense])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"finance_{user['username']}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
    return send_file(bio, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ---------- Frontend page (single-file) ----------
@app.route('/')
def index():
    # Page uses Bootstrap 5, Chart.js, Bootstrap Icons
    return """
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>财务管理（单文件）</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root{
  --soft-green: #e6f7ec;     /* 淡绿色 */
  --soft-gold: #fff7e6;      /* 淡淡金色 */
  --soft-red-bg: #fff1f2;    /* 淡红色背景 */
  --card-border: #e4e4e4;
}
body{
  background: var(--soft-red-bg);
  max-width:1100px;margin:16px auto;padding:12px;
  color:#2b2b2b;
}
.small{font-size:0.9rem;color:#555;}
.card{
  margin-bottom:12px;
  background: linear-gradient(180deg, var(--soft-green), var(--soft-gold));
  border: 1px solid var(--card-border);
  box-shadow: 0 1px 6px rgba(0,0,0,0.03);
}
.table-fixed tbody{height:300px;overflow:auto;display:block;}
.btn-primary{
  background: linear-gradient(180deg,#bfeccf,#9fd8b3);
  border-color: rgba(0,0,0,0.06);
}
.btn-success{
  background: linear-gradient(180deg,#dff7e8,#bfeecd);
  border-color: rgba(0,0,0,0.04);
  color: #0b5b34;
}
.btn-danger{
  background: linear-gradient(180deg,#ffdfcf,#ffc69f);
  border-color: rgba(0,0,0,0.04);
  color:#6b1f1f;
}
.btn-outline-primary{
  background: transparent;
  border-color: rgba(0,0,0,0.06);
}
h3 { margin:0; color:#184d2e; }
</style>
</head>
<body>
<div class="container">
  <div class="d-flex justify-content-between align-items-center mb-3">
    <h3><i class="bi bi-wallet2"></i> 单文件财务管理</h3>
    <div id="userArea"></div>
  </div>

  <div id="authCard" class="card p-3">
    <div class="row g-2 align-items-center">
      <div class="col-auto">
        <input id="username" class="form-control" placeholder="用户名">
      </div>
      <div class="col-auto">
        <input id="password" type="password" class="form-control" placeholder="密码">
      </div>
      <div class="col-auto">
        <button class="btn btn-primary" onclick="login()"><i class="bi bi-box-arrow-in-right"></i> 登录</button>
        <button class="btn btn-outline-secondary" onclick="register()"><i class="bi bi-person-plus"></i> 注册</button>
      </div>
      <div class="col-12 small text-muted">示例账户：<span class="badge bg-primary">demo / demo</span></div>
    </div>
  </div>

  <div id="app" style="display:none;">
    <div class="card p-3">
      <div class="row g-2">
        <div class="col-md-6">
          <input id="desc" class="form-control" placeholder="描述">
        </div>
        <div class="col-md-2">
          <input id="amount" class="form-control" placeholder="金额">
        </div>
        <div class="col-md-2">
          <input id="category" class="form-control" placeholder="分类">
        </div>
        <div class="col-md-2 d-flex gap-2">
          <button class="btn btn-success w-100" onclick="addIncome()"><i class="bi bi-plus-circle"></i> 新增收入</button>
          <button class="btn btn-danger w-100" onclick="addExpense()"><i class="bi bi-dash-circle"></i> 新增支出</button>
        </div>
      </div>
      <div class="mt-2 d-flex gap-2">
        <button class="btn btn-outline-secondary" onclick="openAddModal()">高级添加</button>
        <button class="btn btn-outline-primary" onclick="exportFile()"><i class="bi bi-file-earmark-excel"></i> Export Excel</button>
      </div>
    </div>

    <div class="row">
      <div class="col-lg-7">
        <div class="card p-3">
          <h5>交易列表</h5>
          <div class="row g-2 mb-2">
            <div class="col">
              <input id="filterCategory" class="form-control" placeholder="按分类过滤">
            </div>
            <div class="col-auto">
              <select id="filterType" class="form-select">
                <option value="">全部</option><option value="income">收入</option><option value="expense">支出</option>
              </select>
            </div>
            <div class="col-auto">
              <input id="filterMonth" class="form-control" placeholder="YYYY-MM">
            </div>
            <div class="col-auto">
              <button class="btn btn-secondary" onclick="loadTransactions()">刷新</button>
            </div>
          </div>
          <div id="txTable" class="table-responsive"></div>
        </div>
      </div>

      <div class="col-lg-5">
        <div class="card p-3 mb-3">
          <h5>每月统计（柱状图）</h5>
          <canvas id="monthChart" height="220"></canvas>
        </div>
        <div class="card p-3">
          <h5>月度表格</h5>
          <div id="monthlySummary"></div>
        </div>
      </div>
    </div>
  </div>
</div>

<script>
// Simple local session (localStorage stores user under English key "currentUser")
function setUser(u){ if(u) localStorage.setItem('currentUser', JSON.stringify(u)); else localStorage.removeItem('currentUser'); render(); }
function getUser(){ try{return JSON.parse(localStorage.getItem('currentUser'));}catch(e){return null;} }

function render(){
  const ua = document.getElementById('userArea');
  const user = getUser();
  if(user){
    ua.innerHTML = '已登录：' + escapeHtml(user.username) + ' <button class="btn btn-sm btn-outline-secondary" onclick="logout()">登出</button>';
    document.getElementById('authCard').style.display='none';
    document.getElementById('app').style.display='block';
    loadTransactions(); loadMonthly(); loadChart();
  } else {
    ua.innerHTML = '';
    document.getElementById('authCard').style.display='block';
    document.getElementById('app').style.display='none';
  }
}

// Auth
async function register(){
  const username=document.getElementById('username').value.trim();
  const password=document.getElementById('password').value.trim();
  if(!username||!password){ alert('请输入用户名和密码'); return; }
  const r=await fetch('/api/register',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({username,password})});
  const j=await r.json();
  if(!r.ok){ alert(j.error||'注册失败'); return; }
  setUser({user_id:j.user_id, username:j.username});
}
async function login(){
  const username=document.getElementById('username').value.trim();
  const password=document.getElementById('password').value.trim();
  if(!username||!password){ alert('请输入用户名和密码'); return; }
  const r=await fetch('/api/login',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({username,password})});
  const j=await r.json();
  if(!r.ok){ alert(j.error||'登录失败'); return; }
  setUser({user_id:j.user_id, username:j.username});
}
function logout(){ setUser(null); }

// Quick add
async function addIncome(){ await quickAdd('income'); }
async function addExpense(){ await quickAdd('expense'); }
async function quickAdd(type){
  const user=getUser(); if(!user){ alert('请先登录'); return; }
  const description=document.getElementById('desc').value.trim();
  const amount=document.getElementById('amount').value.trim();
  const category=document.getElementById('category').value.trim();
  if(!description||!amount){ alert('请填写描述和金额'); return; }
  const r=await fetch('/api/add_transaction',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({user_id:user.user_id,description,amount,type,category})});
  const j=await r.json();
  if(!r.ok){ alert(j.error||'添加失败'); return; }
  document.getElementById('desc').value=''; document.getElementById('amount').value=''; document.getElementById('category').value='';
  loadTransactions(); loadMonthly(); loadChart();
}

// Advanced add (with custom date)
function openAddModal(){
  const user=getUser(); if(!user){ alert('请先登录'); return; }
  const description=prompt('描述'); if(description===null) return;
  const amount=prompt('金额'); if(amount===null) return;
  const type=prompt("类型: 'income' 或 'expense'","expense"); if(type===null) return;
  const category=prompt('分类','');
  const date=prompt('日期（ISO 或 YYYY-MM-DD）', new Date().toISOString());
  (async ()=>{
    const r=await fetch('/api/add_transaction',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({user_id:user.user_id,description,amount,type,category,date})});
    const j=await r.json();
    if(!r.ok){ alert(j.error||'添加失败'); return; }
    loadTransactions(); loadMonthly(); loadChart();
  })();
}

// List load & render
async function loadTransactions(){
  const user=getUser(); if(!user) return;
  let url='/api/get_transactions?user_id=' + user.user_id;
  const category=document.getElementById('filterCategory').value.trim();
  const type=document.getElementById('filterType').value;
  const month=document.getElementById('filterMonth').value.trim();
  if(category) url += '&category=' + encodeURIComponent(category);
  if(type) url += '&type=' + encodeURIComponent(type);
  if(month) { url += '&start_date=' + encodeURIComponent(month + '-01T00:00:00') + '&end_date=' + encodeURIComponent(month + '-31T23:59:59'); }
  const r=await fetch(url);
  const data=await r.json();
  if(!r.ok){ alert(data.error||'加载失败'); return; }
  renderTxTable(data);
}

function renderTxTable(txs){
  const c=document.getElementById('txTable');
  if(!txs||txs.length===0){ c.innerHTML='<div class="small">暂无交易</div>'; return; }
  let html='<table class="table table-sm table-hover"><thead><tr><th>日期</th><th>描述</th><th>分类</th><th>类型</th><th class="text-end">金额</th><th>操作</th></tr></thead><tbody>';
  txs.forEach(t=>{
    const amt = Number(t.amount).toFixed(2);
    html += `<tr><td>${escapeHtml(t.date)}</td><td>${escapeHtml(t.description)}</td><td>${escapeHtml(t.category||'')}</td><td>${escapeHtml(t.type)}</td><td class="text-end">${amt}</td>`;
    html += `<td><button class="btn btn-sm btn-outline-primary me-1" onclick="editTx(${t.id})"><i class="bi bi-pencil"></i></button>`;
    html += `<button class="btn btn-sm btn-outline-danger" onclick="deleteTx(${t.id})"><i class="bi bi-trash"></i></button></td></tr>`;
  });
  html += '</tbody></table>';
  c.innerHTML = html;
}

async function deleteTx(id){
  if(!confirm('确认删除？')) return;
  const user=getUser(); if(!user) return;
  const r=await fetch('/api/delete_transaction',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({user_id:user.user_id, id})});
  const j=await r.json();
  if(!r.ok){ alert(j.error||'删除失败'); return; }
  loadTransactions(); loadMonthly(); loadChart();
}

function editTx(id){
  (async ()=>{
    const user=getUser(); if(!user) return;
    const r=await fetch('/api/get_transactions?user_id=' + user.user_id);
    const txs=await r.json();
    const tx=txs.find(x=>x.id===id);
    if(!tx){ alert('未找到'); return; }
    const description=prompt('描述', tx.description); if(description===null) return;
    const amount=prompt('金额', tx.amount); if(amount===null) return;
    const type=prompt("类型 (income 或 expense)", tx.type); if(type===null) return;
    const category=prompt('分类', tx.category||''); if(category===null) return;
    const date=prompt('日期 (ISO)', tx.date); if(date===null) return;
    const res=await fetch('/api/edit_transaction',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({user_id:user.user_id,id,description,amount,type,category,date})});
    const j=await res.json();
    if(!res.ok){ alert(j.error||'更新失败'); return; }
    loadTransactions(); loadMonthly(); loadChart();
  })();
}

// Monthly summary & chart
let monthChart = null;
async function loadMonthly(){
  const user=getUser(); if(!user) return;
  const r=await fetch('/api/monthly_summary?user_id=' + user.user_id);
  const data=await r.json();
  if(!r.ok){ alert(data.error||'加载失败'); return; }
  const el=document.getElementById('monthlySummary');
  if(!data||data.length===0){ el.innerHTML='<div class="small">暂无月度数据</div>'; return; }
  let html = '<table class="table table-sm"><thead><tr><th>月份</th><th class="text-end">收入</th><th class="text-end">支出</th><th class="text-end">结余</th></tr></thead><tbody>';
  data.forEach(m=>{
    html += `<tr><td>${escapeHtml(m.month)}</td><td class="text-end">${Number(m.income).toFixed(2)}</td><td class="text-end">${Number(m.expense).toFixed(2)}</td><td class="text-end">${Number(m.balance).toFixed(2)}</td></tr>`;
  });
  html += '</tbody></table>';
  el.innerHTML = html;
  return data;
}

async function loadChart(){
  const user=getUser(); if(!user) return;
  const r=await fetch('/api/monthly_summary?user_id=' + user.user_id);
  const data=await r.json();
  if(!r.ok) return;
  const labels = data.map(d=>d.month);
  const incomes = data.map(d=>d.income);
  const expenses = data.map(d=>d.expense);
  const ctx = document.getElementById('monthChart').getContext('2d');
  // Chart config: ensure both income and expense datasets are shown distinctly
  if(monthChart) monthChart.destroy();
  monthChart = new Chart(ctx, {
    type: 'bar',
    data: {
      labels: labels,
      datasets: [
        {
          label: '收入',
          data: incomes,
          backgroundColor: 'rgba(40,167,69,0.7)', // 绿色用于收入
          borderColor: 'rgba(40,167,69,0.9)',
          borderWidth: 1
        },
        {
          label: '支出',
          data: expenses,
          backgroundColor: 'rgba(220,53,69,0.7)', // 红色用于支出
          borderColor: 'rgba(220,53,69,0.9)',
          borderWidth: 1
        }
      ]
    },
    options: {
      responsive: true,
      scales: {
        x: { stacked: false },
        y: { beginAtZero: true }
      },
      interaction: { mode: 'index', intersect: false },
      plugins: {
        tooltip: { enabled: true }
      }
    }
  });
}

// Export file (calls new /api/export route)
function exportFile(){
  const user=getUser(); if(!user) return;
  window.location = '/api/export?user_id=' + user.user_id;
}

function escapeHtml(s){ return s? s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;') : ''; }

// Init
render();
</script>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
"""
if __name__ == '__main__':
    with app.app_context():
        get_db()
    app.run(debug=False)
